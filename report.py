#!/usr/bin/env python3
"""AWS Inspector v2 ECR vulnerability report generator."""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import boto3
import openpyxl
from openpyxl.styles import Font


SEVERITY_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNTRIAGED"]
AGE_BUCKETS = ["< 30 days", "30-60 days", "60-90 days", "> 90 days"]


_LATEST_REPORT_RE = re.compile(
    r"^(?P<account>\d+)-inspector-report-(?P<date>\d{6})-(?P<time>\d{4})-latest\.xlsx$"
)


def parse_report_filename(path):
    """Parse a latest-report filename into (account_id, timestamp_utc).

    Returns None if the filename does not match the pattern.
    """
    basename = os.path.basename(path)
    m = _LATEST_REPORT_RE.match(basename)
    if not m:
        return None
    try:
        ts = datetime.strptime(m.group("date") + m.group("time"), "%y%m%d%H%M")
        ts = ts.replace(tzinfo=timezone.utc)
    except ValueError:
        return None
    return (m.group("account"), ts)


def find_history_reports(search_dir, account_id, max_age_days, now=None):
    """Return sorted list of -latest.xlsx report paths matching account_id within max_age_days."""
    if not os.path.isdir(search_dir):
        return []
    if now is None:
        now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=max_age_days)
    results = []
    for name in os.listdir(search_dir):
        parsed = parse_report_filename(name)
        if parsed is None:
            continue
        acct, ts = parsed
        if acct != account_id:
            continue
        if ts < cutoff:
            continue
        results.append(os.path.join(search_dir, name))
    results.sort()
    return results


def _derive_sheet_name_for_repo(repo, used_names):
    """Replicate the sheet-name truncation/collision logic from write_report."""
    base = repo[:31]
    sheet_name = base
    counter = 2
    while sheet_name in used_names:
        suffix = f"_{counter}"
        sheet_name = base[:31 - len(suffix)] + suffix
        counter += 1
    return sheet_name


def _repos_in_order_from_summary(workbook):
    """Read full repo names from the Repository Summary sheet in written order."""
    if "Repository Summary" not in workbook.sheetnames:
        return []
    ws = workbook["Repository Summary"]
    repos = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if name is None or name == "Total":
            continue
        repos.append(str(name))
    return repos


def _parse_history_date(value):
    """Parse a date cell from a past report. Returns UTC datetime or None."""
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            return None
    return None


def read_history_from_report(path):
    """Read {(repo, title): first_discovered_utc} from a past latest-report xlsx.

    Uses the Repository Summary sheet to learn full repo names, then reproduces
    write_report's sheet-name truncation to locate each per-repo sheet. Column
    indices are looked up by header name (resilient to new columns). Returns an
    empty dict on any error (missing file, unreadable, unexpected format).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}

    result = {}
    try:
        repos = _repos_in_order_from_summary(wb)
        used_sheet_names = set()
        for repo in repos:
            sheet_name = _derive_sheet_name_for_repo(repo, used_sheet_names)
            used_sheet_names.add(sheet_name)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            title_col = headers.get("Title")
            date_col = headers.get("First Discovered")
            if not title_col or not date_col:
                continue
            for row in range(2, ws.max_row + 1):
                title = ws.cell(row, title_col).value
                date_raw = ws.cell(row, date_col).value
                if not title or not date_raw:
                    continue
                parsed_date = _parse_history_date(date_raw)
                if parsed_date is None:
                    continue
                key = (repo, str(title))
                if key not in result or parsed_date < result[key]:
                    result[key] = parsed_date
    finally:
        wb.close()
    return result


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate an Excel report from AWS Inspector v2 ECR findings."
    )
    parser.add_argument(
        "--output", default=None,
        help="Output filename (default: <aws_account_id>-inspector-report-<YYMMDD>-<HHmm>.xlsx)"
    )
    parser.add_argument(
        "--severity", action="append", default=[],
        choices=["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFORMATIONAL"],
        metavar="LEVEL",
        help="Filter by severity (repeatable): CRITICAL, HIGH, MEDIUM, LOW, INFORMATIONAL"
    )
    parser.add_argument(
        "--repo", action="append", default=[],
        metavar="NAME",
        help="Filter to specific ECR repo name (repeatable)"
    )
    parser.add_argument(
        "--status", action="append", default=None,
        choices=["ACTIVE", "SUPPRESSED", "CLOSED"],
        metavar="STATUS",
        help="Filter by finding status (default: ACTIVE). Options: ACTIVE, SUPPRESSED, CLOSED"
    )
    parser.add_argument(
        "--region", default=None,
        help="AWS region (default: from AWS_DEFAULT_REGION env var)"
    )
    parser.add_argument(
        "--skip-latest", action="store_true",
        help="Skip generating the latest-image-only report"
    )
    parser.add_argument(
        "--skip-cleanup", action="store_true",
        help="Skip generating the ECR image cleanup report"
    )
    args = parser.parse_args(argv)
    if args.status is None:
        args.status = ["ACTIVE"]
    return args


def get_account_id(region=None):
    """Fetch the AWS account ID using STS."""
    kwargs = {}
    if region:
        kwargs["region_name"] = region
    sts = boto3.client("sts", **kwargs)
    try:
        return sts.get_caller_identity()["Account"]
    except Exception:
        return "unknown"


def _get_image_details(finding):
    """Extract ECR image details from a finding's resources."""
    for resource in finding.get("resources", []):
        details = resource.get("details", {}).get("awsEcrContainerImage", {})
        if details.get("repositoryName"):
            return details
    return {}


def fetch_ecr_images(repo_names, region=None):
    """Fetch all image details from ECR for the given repos.

    Returns {repo: [{"digest", "tags", "pushed_at", "last_pulled"}, ...]}.
    """
    kwargs = {}
    if region:
        kwargs["region_name"] = region
    ecr = boto3.client("ecr", **kwargs)
    all_images = {}
    for repo in repo_names:
        try:
            paginator = ecr.get_paginator("describe_images")
            images = []
            for page in paginator.paginate(repositoryName=repo):
                for img in page.get("imageDetails", []):
                    digest = img.get("imageDigest")
                    pushed = img.get("imagePushedAt")
                    if digest and pushed:
                        images.append({
                            "digest": digest,
                            "tags": img.get("imageTags", []),
                            "pushed_at": pushed,
                            "last_pulled": img.get("lastRecordedPullTime"),
                        })
            all_images[repo] = images
        except Exception:
            continue
    return all_images


def latest_digests_from_images(all_images):
    """Derive the latest image digest per repo from fetch_ecr_images output."""
    latest = {}
    for repo, images in all_images.items():
        best = None
        for img in images:
            if best is None or img["pushed_at"] > best["pushed_at"]:
                best = img
        if best:
            latest[repo] = best["digest"]
    return latest


def filter_latest_image_findings(raw_findings, ecr_latest=None):
    """Return only findings for the latest (most recently pushed) image per repo.

    If ecr_latest is provided (dict of repo -> imageDigest), only include
    findings for repos where the latest Inspector image matches the latest
    ECR image.  A mismatch means the actual latest ECR image has no findings.
    """
    # Track the latest pushedAt per repo
    latest_by_repo = {}  # repo -> (pushedAt, imageHash)
    for f in raw_findings:
        details = _get_image_details(f)
        repo = details.get("repositoryName")
        image_hash = details.get("imageHash")
        pushed_at = details.get("pushedAt")
        if not repo or not image_hash or not pushed_at:
            continue
        if isinstance(pushed_at, str):
            pushed_at = datetime.fromisoformat(pushed_at.replace("Z", "+00:00"))
        current = latest_by_repo.get(repo)
        if current is None or pushed_at > current[0]:
            latest_by_repo[repo] = (pushed_at, image_hash)

    # Exclude repos where Inspector's latest image doesn't match ECR's actual latest
    if ecr_latest:
        for repo in list(latest_by_repo.keys()):
            ecr_digest = ecr_latest.get(repo)
            if ecr_digest and latest_by_repo[repo][1] != ecr_digest:
                del latest_by_repo[repo]

    # Filter findings to only those matching the latest image per repo
    result = []
    for f in raw_findings:
        details = _get_image_details(f)
        repo = details.get("repositoryName")
        image_hash = details.get("imageHash")
        if repo in latest_by_repo and latest_by_repo[repo][1] == image_hash:
            result.append(f)
    return result


def fetch_findings(args):
    """Fetch ECR findings from AWS Inspector v2, returning a flat list."""
    kwargs = {}
    if args.region:
        kwargs["region_name"] = args.region

    client = boto3.client("inspector2", **kwargs)
    paginator = client.get_paginator("list_findings")

    filter_criteria = {
        "findingType": [{"comparison": "EQUALS", "value": "PACKAGE_VULNERABILITY"}],
        "resourceType": [{"comparison": "EQUALS", "value": "AWS_ECR_CONTAINER_IMAGE"}],
        "findingStatus": [
            {"comparison": "EQUALS", "value": s} for s in args.status
        ],
    }

    if args.severity:
        filter_criteria["severity"] = [
            {"comparison": "EQUALS", "value": s} for s in args.severity
        ]

    # Pass repo filter only for single repo (API uses AND for multiple values).
    # For multiple repos, filter client-side after fetching.
    if len(args.repo) == 1:
        filter_criteria["ecrImageRepositoryName"] = [
            {"comparison": "EQUALS", "value": args.repo[0]}
        ]

    findings = []
    for page in paginator.paginate(filterCriteria=filter_criteria):
        findings.extend(page.get("findings", []))

    # Client-side repo filter when multiple repos are specified
    if len(args.repo) > 1:
        repo_set = set(args.repo)
        findings = [
            f for f in findings
            if any(
                resource.get("details", {}).get("awsEcrContainerImage", {}).get("repositoryName") in repo_set
                for resource in f.get("resources", [])
            )
        ]

    return findings


def age_bucket(days: int) -> str:
    """Return the age bucket label for a given number of days."""
    if days < 30:
        return "< 30 days"
    elif days <= 60:
        return "30-60 days"
    elif days <= 90:
        return "60-90 days"
    else:
        return "> 90 days"


def normalize_findings(raw_findings: list) -> list:
    """Normalize raw Inspector findings into a flat list of dicts."""
    now = datetime.now(timezone.utc)
    results = []
    for f in raw_findings:
        severity_raw = f.get("severity", {})
        if isinstance(severity_raw, dict):
            severity_label = severity_raw.get("label", "")
        else:
            severity_label = str(severity_raw) if severity_raw else ""
        if severity_label not in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
            severity_label = "UNTRIAGED"

        first_observed_raw = f.get("firstObservedAt", "")
        try:
            if isinstance(first_observed_raw, datetime):
                first_observed = first_observed_raw
                if first_observed.tzinfo is None:
                    first_observed = first_observed.replace(tzinfo=timezone.utc)
            else:
                first_observed = datetime.fromisoformat(
                    str(first_observed_raw).replace("Z", "+00:00")
                )
        except (ValueError, AttributeError):
            first_observed = now

        days_old = (now - first_observed).days
        vulnerable_packages = f.get("packageVulnerabilityDetails", {}).get("vulnerablePackages", [])
        has_fixed_version = any(pkg.get("fixedInVersion") for pkg in vulnerable_packages)
        remediation_parts = []
        if has_fixed_version:
            remediation_parts.append(
                "Upgrade your installed software packages to the proposed fixed in version and release."
            )
            for pkg in vulnerable_packages:
                pkg_fix = (pkg.get("remediation", "") or "").strip()
                if pkg_fix and pkg_fix not in remediation_parts:
                    remediation_parts.append(pkg_fix)
        else:
            rec_text = (f.get("remediation", {}).get("recommendation", {}).get("text", "") or "").strip()
            if rec_text and rec_text.lower() != "none provided":
                remediation_parts.append(rec_text)
        remediation = "\n".join(remediation_parts)

        repo = ""
        for resource in f.get("resources", []):
            details = resource.get("details", {}).get("awsEcrContainerImage", {})
            if details.get("repositoryName"):
                repo = details["repositoryName"]
                break
        if not repo:
            repo = "(unknown)"

        vulnerability_id = (
            f.get("packageVulnerabilityDetails", {}).get("vulnerabilityId", "") or ""
        )

        results.append({
            "repo": repo,
            "severity": severity_label,
            "title": f.get("title", ""),
            "remediation": remediation,
            "first_observed": first_observed,
            "age_days": days_old,
            "age_bucket": age_bucket(days_old),
            "vulnerability_id": vulnerability_id,
        })
    return results


def build_severity_summary(findings: list) -> dict:
    """Build severity-by-age-bucket summary counts."""
    summary = {
        sev: {"total": 0, **{bucket: 0 for bucket in AGE_BUCKETS}}
        for sev in SEVERITY_ORDER
    }
    for f in findings:
        sev = f["severity"]
        summary[sev]["total"] += 1
        summary[sev][f["age_bucket"]] += 1
    return summary


def build_repo_summary(findings: list) -> dict:
    """Build per-repository severity counts."""
    summary = {}
    for f in findings:
        repo = f["repo"]
        if repo not in summary:
            summary[repo] = {sev: 0 for sev in SEVERITY_ORDER}
            summary[repo]["total"] = 0
        summary[repo][f["severity"]] += 1
        summary[repo]["total"] += 1
    return summary


def build_repo_findings(findings: list) -> dict:
    """Group findings by repository, sorted by severity then first observed date."""
    severity_rank = {sev: i for i, sev in enumerate(SEVERITY_ORDER)}
    grouped = {}
    for f in findings:
        repo = f["repo"]
        grouped.setdefault(repo, []).append(f)
    for repo in grouped:
        grouped[repo].sort(
            key=lambda item: (severity_rank.get(item["severity"], 99), item["first_observed"])
        )
    return grouped


def _bold(ws, row, col):
    ws.cell(row, col).font = Font(bold=True)


def write_report(
    output_path: str,
    severity_summary: dict,
    repo_summary: dict,
    repo_findings: dict,
):
    """Write all report data to an Excel workbook."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Severity Summary ---
    ws1 = wb.active
    ws1.title = "Severity Summary"
    headers = ["Severity Level", "Total"] + AGE_BUCKETS
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        _bold(ws1, 1, col)

    severity_display = {
        "CRITICAL": "Critical", "HIGH": "High", "MEDIUM": "Medium",
        "LOW": "Low", "UNTRIAGED": "Untriaged"
    }
    totals = {bucket: 0 for bucket in ["total"] + AGE_BUCKETS}
    for sev in SEVERITY_ORDER:
        data = severity_summary[sev]
        row = [severity_display[sev], data["total"]] + [data[b] for b in AGE_BUCKETS]
        ws1.append(row)
        totals["total"] += data["total"]
        for b in AGE_BUCKETS:
            totals[b] += data[b]

    total_row = ["Total", totals["total"]] + [totals[b] for b in AGE_BUCKETS]
    ws1.append(total_row)
    for col in range(1, len(headers) + 1):
        _bold(ws1, ws1.max_row, col)

    ws1.freeze_panes = "A2"

    # --- Sheet 2: Repository Summary ---
    ws2 = wb.create_sheet("Repository Summary")
    sev_labels = ["Critical", "High", "Medium", "Low", "Untriaged"]
    sev_keys = SEVERITY_ORDER
    headers2 = ["Amazon ECR Container"] + sev_labels + ["Total"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        _bold(ws2, 1, col)

    repo_totals = {key: 0 for key in sev_keys + ["total"]}
    for repo in sorted(repo_summary.keys()):
        data = repo_summary[repo]
        row = [repo] + [data[k] for k in sev_keys] + [data["total"]]
        ws2.append(row)
        for k in sev_keys:
            repo_totals[k] += data[k]
        repo_totals["total"] += data["total"]

    total_row2 = ["Total"] + [repo_totals[k] for k in sev_keys] + [repo_totals["total"]]
    ws2.append(total_row2)
    for col in range(1, len(headers2) + 1):
        _bold(ws2, ws2.max_row, col)

    ws2.freeze_panes = "A2"

    # --- Sheet 3+: Per-repository findings ---
    used_sheet_names = set()
    for repo in sorted(repo_findings.keys()):
        base = repo[:31]
        sheet_name = base
        counter = 2
        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            sheet_name = base[:31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        repo_headers = ["S/N", "Title", "Remediation", "Severity", "First Discovered", "Vulnerability ID"]
        ws.append(repo_headers)
        for col in range(1, len(repo_headers) + 1):
            _bold(ws, 1, col)
        for i, f in enumerate(repo_findings[repo], start=1):
            ws.append([
                i,
                f["title"],
                f["remediation"],
                f["severity"].capitalize(),
                f["first_observed"].strftime("%Y-%m-%d"),
                f.get("vulnerability_id", ""),
            ])
        ws.freeze_panes = "A2"

    wb.save(output_path)


def write_ecr_cleanup_report(output_path, all_images, latest_digests, region=None):
    """Write an ECR image cleanup report with images to delete and latest images."""
    wb = openpyxl.Workbook()

    region_flag = f" --region {region}" if region else ""

    # --- Sheet 1: Images to Delete ---
    ws1 = wb.active
    ws1.title = "Images to Delete"
    headers = ["Repository", "Image Tags", "Image Digest", "Date Pushed",
               "Date Last Pulled", "Delete Command"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        _bold(ws1, 1, col)

    for repo in sorted(all_images.keys()):
        latest_digest = latest_digests.get(repo)
        old_images = [
            img for img in all_images[repo]
            if img["digest"] != latest_digest
        ]
        old_images.sort(key=lambda img: img["pushed_at"], reverse=True)
        for img in old_images:
            tags = ", ".join(img["tags"]) if img["tags"] else "(untagged)"
            pushed_str = img["pushed_at"].strftime("%Y-%m-%d %H:%M:%S") if img["pushed_at"] else ""
            pulled_str = img["last_pulled"].strftime("%Y-%m-%d %H:%M:%S") if img["last_pulled"] else ""
            cmd = (f"aws ecr batch-delete-image --repository-name {repo}"
                   f" --image-ids imageDigest={img['digest']}{region_flag}")
            ws1.append([repo, tags, img["digest"], pushed_str, pulled_str, cmd])

    ws1.freeze_panes = "A2"

    # --- Sheet 2: Latest Images ---
    ws2 = wb.create_sheet("Latest Images")
    headers2 = ["Repository", "Image Tags", "Image Digest", "Date Pushed",
                "Date Last Pulled"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        _bold(ws2, 1, col)

    for repo in sorted(all_images.keys()):
        latest_digest = latest_digests.get(repo)
        for img in all_images[repo]:
            if img["digest"] == latest_digest:
                tags = ", ".join(img["tags"]) if img["tags"] else "(untagged)"
                pushed_str = img["pushed_at"].strftime("%Y-%m-%d %H:%M:%S") if img["pushed_at"] else ""
                pulled_str = img["last_pulled"].strftime("%Y-%m-%d %H:%M:%S") if img["last_pulled"] else ""
                ws2.append([repo, tags, img["digest"], pushed_str, pulled_str])
                break

    ws2.freeze_panes = "A2"

    wb.save(output_path)


def main():
    args = parse_args()

    if args.output is None:
        account_id = get_account_id(args.region)
        timestamp = datetime.now().strftime("%y%m%d-%H%M")
        args.output = f"{account_id}-inspector-report-{timestamp}.xlsx"

    print("Fetching findings from AWS Inspector v2...")
    raw_findings = fetch_findings(args)
    print(f"Retrieved {len(raw_findings)} findings.")

    findings = normalize_findings(raw_findings)

    severity_summary = build_severity_summary(findings)
    repo_summary = build_repo_summary(findings)
    repo_findings = build_repo_findings(findings)

    write_report(args.output, severity_summary, repo_summary, repo_findings)
    print(f"Report written to: {args.output}")

    need_ecr = not args.skip_latest or not args.skip_cleanup
    ecr_images = {}
    ecr_latest = {}
    if need_ecr:
        repos_in_findings = set()
        for f in raw_findings:
            details = _get_image_details(f)
            repo = details.get("repositoryName")
            if repo:
                repos_in_findings.add(repo)
        print("Querying ECR for image details...")
        ecr_images = fetch_ecr_images(repos_in_findings, args.region)
        ecr_latest = latest_digests_from_images(ecr_images)

    if not args.skip_latest:
        latest_raw = filter_latest_image_findings(raw_findings, ecr_latest)
        print(f"Filtered to {len(latest_raw)} findings for latest images.")
        latest_findings = normalize_findings(latest_raw)
        latest_severity = build_severity_summary(latest_findings)
        latest_repo = build_repo_summary(latest_findings)
        latest_repo_findings = build_repo_findings(latest_findings)
        latest_output = args.output.replace(".xlsx", "-latest.xlsx")
        write_report(latest_output, latest_severity, latest_repo, latest_repo_findings)
        print(f"Latest-image report written to: {latest_output}")

    if not args.skip_cleanup:
        cleanup_output = args.output.replace(".xlsx", "-ecr-cleanup.xlsx")
        write_ecr_cleanup_report(cleanup_output, ecr_images, ecr_latest, args.region)
        total_old = sum(
            sum(1 for img in imgs if img["digest"] != ecr_latest.get(repo))
            for repo, imgs in ecr_images.items()
        )
        print(f"ECR cleanup report written to: {cleanup_output} ({total_old} old images across {len(ecr_images)} repos)")


if __name__ == "__main__":
    main()
