import os
import sys
from unittest.mock import MagicMock, patch
from datetime import datetime, timezone, timedelta
import pytest
import openpyxl
from report import main


def make_raw_finding(severity="HIGH", repo="my-app", days_ago=10):
    first_observed = datetime.now(timezone.utc) - timedelta(days=days_ago)
    return {
        "findingArn": f"arn:aws:inspector2:us-east-1:123:finding/{repo}-{severity}",
        "title": f"CVE-2023-1234 - {repo}",
        "title": f"Vulnerability in {repo}.",
        "severity": {"label": severity},
        "firstObservedAt": first_observed.isoformat(),
        "status": "ACTIVE",
        "remediation": {"recommendation": {"text": f"Fix {repo}"}},
        "resources": [{
            "type": "AWS_ECR_CONTAINER_IMAGE",
            "details": {"awsEcrContainerImage": {"repositoryName": repo}}
        }],
    }


@patch("report.boto3.client")
def test_main_creates_report(mock_client_factory, tmp_path):
    mock_client = MagicMock()
    mock_client_factory.return_value = mock_client
    paginator = MagicMock()
    mock_client.get_paginator.return_value = paginator
    paginator.paginate.return_value = [{
        "findings": [
            make_raw_finding("CRITICAL", "app-a", 5),
            make_raw_finding("HIGH", "app-a", 100),
            make_raw_finding("MEDIUM", "app-b", 45),
        ]
    }]

    output = str(tmp_path / "integration_report.xlsx")
    sys.argv = ["report.py", "--output", output]
    main()

    assert os.path.exists(output)
    wb = openpyxl.load_workbook(output)
    assert "Severity Summary" in wb.sheetnames
    assert "Repository Summary" in wb.sheetnames
    assert "app-a" in wb.sheetnames
    assert "app-b" in wb.sheetnames


@patch("report.boto3.client")
def test_main_empty_findings_creates_report(mock_client_factory, tmp_path):
    mock_client = MagicMock()
    mock_client_factory.return_value = mock_client
    paginator = MagicMock()
    mock_client.get_paginator.return_value = paginator
    paginator.paginate.return_value = [{"findings": []}]

    output = str(tmp_path / "empty_report.xlsx")
    sys.argv = ["report.py", "--output", output]
    main()

    assert os.path.exists(output)
    wb = openpyxl.load_workbook(output)
    assert "Severity Summary" in wb.sheetnames


def test_latest_report_inherits_first_discovered_from_prior_report(tmp_path, monkeypatch):
    from datetime import datetime, timezone
    from report import (
        normalize_findings, apply_history, load_history,
        build_severity_summary, build_repo_summary, build_repo_findings,
        write_report,
    )

    # Step 1: write a "prior" report with an old date for CVE-A in app-a
    old_date = datetime(2025, 1, 1, tzinfo=timezone.utc)
    prior_findings = [{
        "repo": "app-a",
        "severity": "HIGH",
        "title": "CVE-A",
        "remediation": "",
        "first_observed": old_date,
        "age_days": 400,
        "age_bucket": "> 90 days",
        "vulnerability_id": "CVE-2025-0001",
    }]
    prior_path = str(tmp_path / "111-inspector-report-250101-1200-latest.xlsx")
    write_report(
        prior_path,
        build_severity_summary(prior_findings),
        build_repo_summary(prior_findings),
        build_repo_findings(prior_findings),
    )

    # Step 2: simulate current Inspector findings where the image is brand-new
    new_date = datetime(2026, 4, 1, tzinfo=timezone.utc)
    current = [{
        "repo": "app-a",
        "severity": "HIGH",
        "title": "CVE-A",
        "remediation": "",
        "first_observed": new_date,
        "age_days": 13,
        "age_bucket": "< 30 days",
        "vulnerability_id": "CVE-2025-0001",
    }]

    # Step 3: load history and apply
    history = load_history(str(tmp_path), "111", max_age_days=600,
                           now=datetime(2026, 4, 14, tzinfo=timezone.utc))
    assert history[("app-a", "CVE-A")] == old_date
    matched = apply_history(current, history, now=datetime(2026, 4, 14, tzinfo=timezone.utc))
    assert matched == 1
    assert current[0]["first_observed"] == old_date
    assert current[0]["age_bucket"] == "> 90 days"
