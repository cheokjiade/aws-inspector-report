"""Tests for the CVE ignore-list feature."""
import os
from datetime import datetime, timezone

import openpyxl
import pytest

from report import (
    build_ignore_map,
    parse_ignore_file,
    partition_by_ignore,
    build_severity_summary,
    build_repo_summary,
    build_repo_findings,
    write_report,
    read_history_from_report,
)


class _Args:
    def __init__(self, ignore_cve=None, ignore_file=None):
        self.ignore_cve = ignore_cve or []
        self.ignore_file = ignore_file


def _normalized(vuln_id="CVE-2024-0001", repo="app-a", severity="HIGH", title="Vuln"):
    return {
        "repo": repo,
        "severity": severity,
        "title": title,
        "remediation": "Upgrade.",
        "first_observed": datetime(2024, 6, 1, tzinfo=timezone.utc),
        "age_days": 10,
        "age_bucket": "< 30 days",
        "vulnerability_id": vuln_id,
    }


# --- parse_ignore_file ---

def test_parse_ignore_file_basic(tmp_path):
    p = tmp_path / "ignore.txt"
    p.write_text("CVE-2024-1111\nCVE-2024-2222\n")
    result = parse_ignore_file(str(p))
    assert result == {"CVE-2024-1111": "", "CVE-2024-2222": ""}


def test_parse_ignore_file_with_reasons_and_comments(tmp_path):
    p = tmp_path / "ignore.txt"
    p.write_text(
        "# top-level comment\n"
        "\n"
        "CVE-2024-1111  # accepted risk - mitigated by WAF\n"
        "CVE-2024-2222\n"
    )
    result = parse_ignore_file(str(p))
    assert result == {
        "CVE-2024-1111": "accepted risk - mitigated by WAF",
        "CVE-2024-2222": "",
    }


# --- build_ignore_map ---

def test_build_ignore_map_combines_cli_and_file(tmp_path):
    p = tmp_path / "ignore.txt"
    p.write_text("CVE-2024-FILE  # from file\n")
    args = _Args(ignore_cve=["CVE-2024-CLI"], ignore_file=str(p))
    result = build_ignore_map(args)
    assert result == {"CVE-2024-FILE": "from file", "CVE-2024-CLI": ""}


def test_build_ignore_map_empty_when_no_input():
    assert build_ignore_map(_Args()) == {}


# --- partition_by_ignore ---

def test_partition_separates_ignored_findings():
    findings = [
        _normalized("CVE-2024-0001"),
        _normalized("CVE-2024-9999"),
    ]
    active, ignored = partition_by_ignore(findings, {"CVE-2024-9999": "false positive"})
    assert len(active) == 1
    assert active[0]["vulnerability_id"] == "CVE-2024-0001"
    assert len(ignored) == 1
    assert ignored[0]["vulnerability_id"] == "CVE-2024-9999"
    assert ignored[0]["ignore_reason"] == "false positive"


def test_partition_case_insensitive():
    findings = [_normalized("CVE-2024-0001")]
    active, ignored = partition_by_ignore(findings, {"cve-2024-0001": ""})
    assert len(active) == 0
    assert len(ignored) == 1


def test_partition_empty_map_returns_all_active():
    findings = [_normalized("CVE-2024-0001")]
    active, ignored = partition_by_ignore(findings, {})
    assert len(active) == 1
    assert ignored == []


def test_partition_does_not_mutate_originals():
    finding = _normalized("CVE-2024-0001")
    findings = [finding]
    _, ignored = partition_by_ignore(findings, {"CVE-2024-0001": "x"})
    assert "ignore_reason" not in finding
    assert ignored[0]["ignore_reason"] == "x"


def test_partition_finding_without_vuln_id_stays_active():
    findings = [_normalized(vuln_id="")]
    active, ignored = partition_by_ignore(findings, {"CVE-2024-0001": ""})
    assert len(active) == 1
    assert ignored == []


# --- write_report ignored sheet ---

def test_write_report_creates_ignored_sheet(tmp_path):
    output = str(tmp_path / "r.xlsx")
    active = [_normalized("CVE-2024-0001", repo="app-a")]
    ignored = [
        {**_normalized("CVE-2024-9999", repo="app-b", title="Ignored vuln"),
         "ignore_reason": "accepted risk"}
    ]
    write_report(
        output,
        build_severity_summary(active),
        build_repo_summary(active),
        build_repo_findings(active),
        ignored_findings=ignored,
    )
    wb = openpyxl.load_workbook(output)
    assert "Ignored Findings" in wb.sheetnames
    ws = wb["Ignored Findings"]
    assert [ws.cell(1, c).value for c in range(1, 8)] == [
        "S/N", "Repository", "Title", "Severity",
        "First Discovered", "Vulnerability ID", "Ignore Reason",
    ]
    assert ws.cell(2, 2).value == "app-b"
    assert ws.cell(2, 3).value == "Ignored vuln"
    assert ws.cell(2, 6).value == "CVE-2024-9999"
    assert ws.cell(2, 7).value == "accepted risk"


def test_write_report_omits_ignored_sheet_when_no_ignored(tmp_path):
    output = str(tmp_path / "r.xlsx")
    active = [_normalized("CVE-2024-0001", repo="app-a")]
    write_report(
        output,
        build_severity_summary(active),
        build_repo_summary(active),
        build_repo_findings(active),
    )
    wb = openpyxl.load_workbook(output)
    assert "Ignored Findings" not in wb.sheetnames


# --- history reads ignored sheet ---

def test_read_history_includes_ignored_sheet(tmp_path):
    """An ignored CVE's first-discovered date is preserved across reports."""
    output = str(tmp_path / "111111111111-inspector-report-240101-1200-latest.xlsx")
    active = [_normalized("CVE-2024-0001", repo="app-a", title="Active")]
    old_date = datetime(2023, 1, 15, tzinfo=timezone.utc)
    ignored = [{
        **_normalized("CVE-2024-9999", repo="app-b", title="Ignored"),
        "first_observed": old_date,
        "ignore_reason": "accepted",
    }]
    write_report(
        output,
        build_severity_summary(active),
        build_repo_summary(active),
        build_repo_findings(active),
        ignored_findings=ignored,
    )
    history = read_history_from_report(output)
    assert ("app-a", "Active") in history
    assert ("app-b", "Ignored") in history
    assert history[("app-b", "Ignored")].date() == old_date.date()
