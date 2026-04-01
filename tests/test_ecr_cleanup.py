import os
from datetime import datetime, timezone
import openpyxl
from report import write_ecr_cleanup_report


def _img(digest, tags, pushed_at, last_pulled=None):
    return {
        "digest": digest,
        "tags": tags,
        "pushed_at": pushed_at,
        "last_pulled": last_pulled,
    }


def test_cleanup_report_separates_latest_and_old(tmp_path):
    all_images = {
        "app-a": [
            _img("sha256:old1", ["v1"], datetime(2024, 1, 1, tzinfo=timezone.utc),
                 datetime(2024, 2, 1, tzinfo=timezone.utc)),
            _img("sha256:old2", [], datetime(2024, 3, 1, tzinfo=timezone.utc)),
            _img("sha256:latest", ["v3", "latest"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"app-a": "sha256:latest"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)
    assert "Images to Delete" in wb.sheetnames
    assert "Latest Images" in wb.sheetnames

    # Images to Delete should have 2 old images (not the latest)
    ws_delete = wb["Images to Delete"]
    rows = list(ws_delete.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    digests = [r[2] for r in rows]
    assert "sha256:old1" in digests
    assert "sha256:old2" in digests
    assert "sha256:latest" not in digests

    # Latest Images should have 1 row
    ws_latest = wb["Latest Images"]
    latest_rows = list(ws_latest.iter_rows(min_row=2, values_only=True))
    assert len(latest_rows) == 1
    assert latest_rows[0][2] == "sha256:latest"
    assert latest_rows[0][1] == "v3, latest"


def test_cleanup_report_delete_command_format(tmp_path):
    all_images = {
        "my-repo": [
            _img("sha256:old", ["v1"], datetime(2024, 1, 1, tzinfo=timezone.utc)),
            _img("sha256:new", ["v2"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"my-repo": "sha256:new"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests, region="us-east-1")

    wb = openpyxl.load_workbook(output)
    ws = wb["Images to Delete"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cmd = row[5]
    assert "aws ecr batch-delete-image" in cmd
    assert "--repository-name my-repo" in cmd
    assert "--image-ids imageDigest=sha256:old" in cmd
    assert "--region us-east-1" in cmd


def test_cleanup_report_no_region_flag(tmp_path):
    all_images = {
        "my-repo": [
            _img("sha256:old", ["v1"], datetime(2024, 1, 1, tzinfo=timezone.utc)),
            _img("sha256:new", ["v2"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"my-repo": "sha256:new"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)
    ws = wb["Images to Delete"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cmd = row[5]
    assert "--region" not in cmd


def test_cleanup_report_untagged_images(tmp_path):
    all_images = {
        "app-a": [
            _img("sha256:untagged", [], datetime(2024, 1, 1, tzinfo=timezone.utc)),
            _img("sha256:tagged", ["latest"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"app-a": "sha256:tagged"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)
    ws = wb["Images to Delete"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[1] == "(untagged)"


def test_cleanup_report_latest_has_no_delete_command(tmp_path):
    all_images = {
        "app-a": [
            _img("sha256:latest", ["v1"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"app-a": "sha256:latest"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)
    # Latest Images sheet should have no delete command column
    ws = wb["Latest Images"]
    headers = [cell.value for cell in ws[1]]
    assert "Delete Command" not in headers

    # Images to Delete should be empty (only image is the latest)
    ws_delete = wb["Images to Delete"]
    rows = list(ws_delete.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 0


def test_cleanup_report_multiple_repos(tmp_path):
    all_images = {
        "repo-a": [
            _img("sha256:a_old", ["v1"], datetime(2024, 1, 1, tzinfo=timezone.utc)),
            _img("sha256:a_new", ["v2"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
        "repo-b": [
            _img("sha256:b_old", ["v1"], datetime(2024, 2, 1, tzinfo=timezone.utc)),
            _img("sha256:b_new", ["v2"], datetime(2024, 7, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"repo-a": "sha256:a_new", "repo-b": "sha256:b_new"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)

    ws_delete = wb["Images to Delete"]
    rows = list(ws_delete.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    repos = [r[0] for r in rows]
    assert "repo-a" in repos
    assert "repo-b" in repos

    ws_latest = wb["Latest Images"]
    latest_rows = list(ws_latest.iter_rows(min_row=2, values_only=True))
    assert len(latest_rows) == 2


def test_cleanup_report_last_pulled_shown(tmp_path):
    pulled = datetime(2024, 5, 15, 10, 30, 0, tzinfo=timezone.utc)
    all_images = {
        "app-a": [
            _img("sha256:old", ["v1"], datetime(2024, 1, 1, tzinfo=timezone.utc), pulled),
            _img("sha256:new", ["v2"], datetime(2024, 6, 1, tzinfo=timezone.utc)),
        ],
    }
    latest_digests = {"app-a": "sha256:new"}
    output = str(tmp_path / "cleanup.xlsx")
    write_ecr_cleanup_report(output, all_images, latest_digests)

    wb = openpyxl.load_workbook(output)
    ws = wb["Images to Delete"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[4] == "2024-05-15 10:30:00"
