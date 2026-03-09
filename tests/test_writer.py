import os
import tempfile
from datetime import datetime, timezone
import pytest
import openpyxl
from report import write_report, build_severity_summary, build_repo_summary, build_repo_findings


def make_normalized(severity="HIGH", repo="my-app", age_bucket="< 30 days"):
    return {
        "repo": repo,
        "severity": severity,
        "title": "A vulnerability.",
        "remediation": "Update the package.",
        "first_observed": datetime(2024, 6, 1, tzinfo=timezone.utc),
        "age_days": 10,
        "age_bucket": age_bucket,
    }


@pytest.fixture
def sample_findings():
    return [
        make_normalized("CRITICAL", "app-a", "< 30 days"),
        make_normalized("HIGH", "app-a", "> 90 days"),
        make_normalized("HIGH", "app-b", "30-60 days"),
    ]


@pytest.fixture
def report_file(sample_findings, tmp_path):
    output = str(tmp_path / "test_report.xlsx")
    severity_summary = build_severity_summary(sample_findings)
    repo_summary = build_repo_summary(sample_findings)
    repo_findings = build_repo_findings(sample_findings)
    write_report(output, severity_summary, repo_summary, repo_findings)
    return output


def test_creates_file(report_file):
    assert os.path.exists(report_file)


def test_sheet_names(report_file):
    wb = openpyxl.load_workbook(report_file)
    sheets = wb.sheetnames
    assert sheets[0] == "Severity Summary"
    assert sheets[1] == "Repository Summary"
    assert "app-a" in sheets
    assert "app-b" in sheets


def test_severity_summary_headers(report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["Severity Summary"]
    headers = [ws.cell(1, c).value for c in range(1, 7)]
    assert headers[0] == "Severity Level"
    assert headers[1] == "Total"
    assert "< 30 days" in headers


def test_severity_summary_values(report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["Severity Summary"]
    # Find CRITICAL row (row 2)
    row_vals = [ws.cell(2, c).value for c in range(1, 7)]
    assert row_vals[0] == "Critical"
    assert row_vals[1] == 1  # total CRITICAL


def test_repo_summary_headers(report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["Repository Summary"]
    headers = [ws.cell(1, c).value for c in range(1, 8)]
    assert headers[0] == "Amazon ECR Container"
    assert "Critical" in headers
    assert "Total" in headers


def test_per_repo_sheet_headers(report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["app-a"]
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == ["S/N", "Title", "Remediation", "Severity", "First Discovered"]


def test_per_repo_sheet_row_count(report_file):
    wb = openpyxl.load_workbook(report_file)
    ws = wb["app-a"]
    # 1 header + 2 findings
    assert ws.max_row == 3


def test_sheet_name_truncated_to_31_chars(tmp_path):
    long_repo = "a" * 40
    findings = [make_normalized("HIGH", long_repo)]
    output = str(tmp_path / "long.xlsx")
    write_report(
        output,
        build_severity_summary(findings),
        build_repo_summary(findings),
        build_repo_findings(findings),
    )
    wb = openpyxl.load_workbook(output)
    for sheet in wb.sheetnames[2:]:
        assert len(sheet) <= 31


def test_empty_repo_name_uses_unknown(tmp_path):
    findings = [{
        "repo": "(unknown)",
        "severity": "HIGH",
        "title": "A vulnerability.",
        "remediation": "",
        "first_observed": datetime(2024, 6, 1, tzinfo=timezone.utc),
        "age_days": 10,
        "age_bucket": "< 30 days",
    }]
    output = str(tmp_path / "unknown_repo.xlsx")
    write_report(
        output,
        build_severity_summary(findings),
        build_repo_summary(findings),
        build_repo_findings(findings),
    )
    wb = openpyxl.load_workbook(output)
    assert "(unknown)" in wb.sheetnames


def test_duplicate_truncated_sheet_names_no_crash(tmp_path):
    findings = [
        make_normalized("HIGH", "a" * 32 + "X"),
        make_normalized("HIGH", "a" * 32 + "Y"),
    ]
    output = str(tmp_path / "dup_names.xlsx")
    write_report(
        output,
        build_severity_summary(findings),
        build_repo_summary(findings),
        build_repo_findings(findings),
    )
    wb = openpyxl.load_workbook(output)
    # Should have 2 distinct per-repo sheets, both <= 31 chars
    repo_sheets = wb.sheetnames[2:]
    assert len(repo_sheets) == 2
    assert len(set(repo_sheets)) == 2  # no duplicates
    for s in repo_sheets:
        assert len(s) <= 31
